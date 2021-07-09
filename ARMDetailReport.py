import cx_Oracle
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import openpyxl
import os
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from cryptography.fernet import Fernet
import pandas as pd
import datetime
from pandas.io.common import EmptyDataError
import logging

# logger definition
def get_logger(script, logname, loglevel="DEBUG"):
    # Default the logging level to INFO if no LOGGING_LEVEL param defined
    level = os.environ.get("ARM_REPORT_DETAILS")
    if level is None:
        level = loglevel
    # Create/Get a custom logger
    logger = logging.getLogger(script, )
    logger.setLevel(logging.DEBUG)

    fh = logging.FileHandler(logname)
    # Set the level
    if level:
        fh.setLevel(logging.getLevelName(level))
    else:
        fh.setLevel(logging.DEBUG)

    # create console handler with a higher log level
    ch = logging.StreamHandler()
    # ch.setLevel(logging.ERROR)
    ch.setLevel(logging.DEBUG)

    # Create formatter and add it to handler
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    fh.setFormatter(formatter)
    ch.setFormatter(formatter)

    # add the handlers to the logger
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger
logger = get_logger('ARM_REPORT_DETAILS.py', 'ARM_REPORT_DETAILS.log')
print = logger.debug


# Change current working dir
wrk_dir=r'E:\Control-M\tabdev\FPM\Tableau_Python_ARM_Detail_Report\\'
os.chdir(wrk_dir)

# Archive folder
if not os.path.exists('Archive'):
    os.makedirs('Archive')

# function definitions
def fileconversion(inputfile):
    df = pd.read_csv(inputfile, dtype='unicode')
    cols = df.columns.tolist()
    cols.remove('Measure Values')
    df = df.set_index(cols, drop=True).unstack('Measure Names').reset_index()
    df.columns = [''.join(col).strip().replace('Measure Values', '') for col in df.columns.values]
    df['Balance Amount'] = df['Balance Amount'].str.replace(",", "")
    for i in ['Balance Amount','Period Name','Reconciliation ID','Rejects']:
        df[i]=df[i].astype(float)
    df['End Date'] = pd.to_datetime(df['End Date'])
    df['Actual End Date'] = pd.to_datetime(df['Actual End Date'])
    df['BU Name']  = df['Reconciliation Name'].str.split("-", n=-1, expand=True)[0]
    cols=['Period Name','Status','Timeliness','Format Name','Reconciliation ID','Reconciliation Name','End Date','Actual End Date','Balance Buckets','Balance Amount','Preparer','Reviewer','Auto Reconciled','User Region','User Region (Reviewer)','Legal Entity','BU Name','Account','Department','Department (Reviewer)','Reconciliation Account ID','GPO Leader','GPO Leader (Reviewer)','Auto Closed','Rejects']
    # xls_writer = pd.ExcelWriter("output.xlsx", engine='xlsxwriter', datetime_format='dd-mmmm-yyyy',options={'strings_to_numbers': True})
    xls_writer = pd.ExcelWriter("output.xlsx", engine='xlsxwriter', datetime_format='dd-mmmm-yyyy')
    df.to_excel(xls_writer, sheet_name=filtervalue, float_format="%.2f", index=False,columns=cols)
    xls_writer.save()

    wb = openpyxl.load_workbook("output.xlsx")
    worksheet = wb.active
    worksheet.title = filtervalue

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))


    for col in worksheet.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            cell.font= Font(name='Tahoma',size=8,color='00000000')
            cell.border = thin_border
            # if column in ['J', 'L', 'M']:
            #     cell.alignment = Alignment(horizontal='left')
            #     cell.value = str(cell.value)

            if cell.coordinate in worksheet.merged_cells:  # not check merge_cells
                continue
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length+0.5)
        worksheet.column_dimensions[column].width = adjusted_width


    for cell in worksheet["1:1"]:
        cell.font = Font(name='Tahoma',size=8,color='00000000', bold=True)
        cell.fill = PatternFill(start_color="BFD2E2", end_color="FFC7CE", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    c = worksheet['A2']
    worksheet.freeze_panes = c
    global filename
    filename='ARM_Detail_Report_'+filtervalue + ".xlsx"
    wb.save(filename)
    wb.close()
    os.remove("output.xlsx")
    os.remove(inputfile)
    return 'Y'
def sendErrormail(err):
    msg = MIMEMultipart()
    msg['From'] = from_id
    msg['To'] = fail_mailid
    msg['Subject'] = 'Action Required: Delivery Failed - ARM Detail Report'
    message = 'Error Occurred : '+err
    msg.attach(MIMEText(message, 'plain'))

    try:
        s.send_message(msg)
        print("Error Occurred Email successfully sent to Internal team")
    except:
        print("Mail Server Error - Error Occurred Email not sent to Internal team")
    del msg
def sendFilemail():
    msg = MIMEMultipart()
    msg['From'] = from_id
    msg['To'] = user_mailid
    msg['Cc'] = user_ccid
    msg['Bcc'] = bcc_id
    msg['Subject'] = 'ARM Detail Report - '+filtervalue
    message = 'Please find the attached ARM Detail Report for Period ' + filtervalue+'.'
    msg.attach(MIMEText(message, 'plain'))

    mail_attach = MIMEApplication(open(filename, "rb").read())
    mail_attach.add_header('Content-Disposition', 'attachment',
                           filename=filename)
    msg.attach(mail_attach)

    try:
        s.send_message(msg)
        print("Email successfully sent to users")
    except:
        print("Mail Server Error - Email not sent to users")
    del msg
def read_csvfile(file):
    try:
        df = pd.read_csv(file, dtype='unicode')
    except EmptyDataError:
        df = pd.DataFrame()
    return df


# reading Bursting_Details_ARM
wb = openpyxl.load_workbook('Bursting_Details_ARM.xlsx')

# mail-server details:
sheet = wb["Server&LoginDetails"]
host = sheet['C4'].value
port = sheet['C5'].value
from_id = sheet['C6'].value
user_mailid = sheet['C8'].value
user_ccid = sheet['C9'].value
bcc_id = sheet['C7'].value
fail_mailid = sheet['C10'].value

#database details:
sheet = wb["DatabaseDetails"]
dbuserid = sheet['B1'].value
login_encrypt_password = sheet['B2'].value
dbhost = sheet['B3'].value
dbport = sheet['B4'].value
dbSID = sheet['B5'].value
dbtns=cx_Oracle.makedsn(dbhost,dbport,dbSID)
key=b'icX0uAcHik9UCEgzTY3jP2_KhbEfXGAZucdSX3sbQMQ='
cipher_suite = Fernet(key)
dbpassword = cipher_suite.decrypt(bytes(login_encrypt_password,'utf-8')).decode('utf-8')





if __name__ == '__main__':
    print ('Script Execution Started')
    # Set Default value for RunFlag, FileFlag, filter value
    RunFlag = 'N'
    FileFlag = 'N'
    filtervalue=''

    # Start Email
    s = smtplib.SMTP(host=host, port=port)
    s.starttls()

    # Override RunFlag
    try:
        connection = cx_Oracle.connect(dbuserid,dbpassword,dbtns)
        cursor = connection.cursor()
        querystring = '''
        select nk_month_id from d_month where LOAD_FIS_MONTH_FLG = 'Y'
        and TRUNC(SYSDATE) = TRUNC(UDE_GPM_DAY_1_DT+1)
         '''
        cursor.execute(querystring)
        row=cursor.fetchone()
        if row is not None:
            RunFlag = 'Y'
            filtervalue=row[0]
        cursor.close ()
        connection.close ()
    except:
        print('Error Occurred while connecting to database')
        sendErrormail('Error Occurred while connecting to database')
    print('RunFlag - ' + RunFlag)
    print ('Period - ' + filtervalue)

    # hardcoding for testing
    # RunFlag = 'Y'
    # filtervalue='202103'

    if RunFlag =='Y':
        print('File Generation Process started ...')
        # reading Workbook & Dashboard name
        sheet = wb["DashboardNameCoded"]
        workbook = sheet['C2'].value
        listreport = sheet['A2'].value
        # server & log in details:
        sheet = wb["Server&LoginDetails"]
        server = sheet['C1'].value
        login_id = sheet['C2'].value
        login_encrypt_password = sheet['C3'].value
        key = b'icX0uAcHik9UCEgzTY3jP2_KhbEfXGAZucdSX3sbQMQ='
        cipher_suite = Fernet(key)
        login_password = cipher_suite.decrypt(bytes(login_encrypt_password, 'utf-8')).decode('utf-8')
        cmd1 = 'tabcmd login -s ' + server + ' -t FPM -u ' + login_id + ' -p ' + login_password + ' --no-certcheck'
        print(os.system(cmd1))

        # file generation Process
        url = '/views/' + workbook + '/' + listreport + '.csv?Period%20Name=' + filtervalue
        output_csv = 'ARM_Detail_Report.csv'
        cmd2 = r'tabcmd get "' + url + '" -f "' + output_csv + '"'
        print(os.system(cmd2))
        cmd3 = r' tabcmd logout'
        print(os.system(cmd3))


        try:
            FileFlag = fileconversion(output_csv)
            print('FileFlag - ' + FileFlag)
        except EmptyDataError:
            print('No Data Available for ' + filtervalue)
            sendErrormail('No Data Available for ' + filtervalue)
        except Exception as e:
            print ('Error Occurred while processing file - '+ str(e))
            sendErrormail('Error Occurred while processing file - '+ str(e))


        if FileFlag == 'Y':
            sendFilemail()
            timestamp = datetime.datetime.today().strftime('%Y-%m-%d_%H_%M_%S')
            os.rename(filename, 'Archive\\' + filename[:-5]+'_'+timestamp+'.xlsx')

    else:
        print('Today is not Day 4 - No File Generation')
    s.quit()
    print ('Script Execution Ended')

