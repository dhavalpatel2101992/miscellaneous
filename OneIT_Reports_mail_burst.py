import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from PyPDF2 import PdfFileReader, PdfFileWriter
import time
import openpyxl
import csv
import os
from pyexcel.cookbook import merge_all_to_a_book
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
import glob
from openpyxl.styles import PatternFill
import pandas as pd
import datetime

# Change current working dir
wrk_dir=r'E:\Control-M\tabprd\FPM\Tableau_Python_OneIT_Timetracking\\'
os.chdir(wrk_dir)

# reading excel
wb = openpyxl.load_workbook('Bursting_Details.xlsx')

# server & log in details:
sheet = wb["Server&LoginDetails"]
server = sheet['C1'].value
login_id = sheet['C2'].value
login_password = sheet['C3'].value

# mail-server details:
host = sheet['C4'].value
port = sheet['C5'].value
from_id = sheet['C6'].value
bcc_id = sheet['C7'].value
subject = sheet['C8'].value

if not os.path.exists('Archive'):
    os.makedirs('Archive')

sheet = wb["BurstingList"]

s = smtplib.SMTP(host=host, port=port)
s.starttls()

for i in range(2, sheet.max_row + 1):
    user_name = sheet['A' + str(i)].value
    user_mailid = sheet['B' + str(i)].value
    user_ccid = sheet['C' + str(i)].value
    user_msg = sheet['D' + str(i)].value
    user_namecoded = user_name.replace(",", "").replace(" ", "%20")
    user_namefilecoded = ''.join(filter(str.isalpha, user_name))
    msg = MIMEMultipart()
    msg['From'] = from_id
    msg['To'] = user_mailid
    msg['Cc'] = user_ccid
    msg['Bcc'] = bcc_id
    msg['Subject'] = subject + " (" + user_name + ")"
    message = user_msg
    msg.attach(MIMEText(message, 'plain'))
    timestamp = datetime.datetime.today().strftime('%Y-%m-%d_%H%M%S')
    if os.path.isfile(user_namefilecoded + '.pdf') and os.path.isfile(user_namefilecoded + '.xlsx'):
        mail_attach = MIMEApplication(open(user_namefilecoded + '.pdf', "rb").read())
        mail_attach.add_header('Content-Disposition', 'attachment',
                               filename="OneIT Labor Group Summary - Week Complete.pdf")
        msg.attach(mail_attach)
        file_pdf = user_namefilecoded + '.pdf'

        mail_attach = MIMEApplication(open(user_namefilecoded + ".xlsx", "rb").read())
        mail_attach.add_header('Content-Disposition', 'attachment',
                               filename="OneIT Labor Employee Detail - Week Complete.xlsx")
        msg.attach(mail_attach)
        file_xlsx = user_namefilecoded + ".xlsx"

        try:
            s.send_message(msg)
            print("Email successfully sent to", user_mailid)
            os.rename(file_pdf, 'Archive\\'+ user_namefilecoded + '_' + timestamp + '.pdf' )
            os.rename(file_xlsx, 'Archive\\'+ user_namefilecoded + '_' + timestamp + '.xlsx' )
        except:
            print("Error - Email not sent to", user_mailid)
        del msg
    else:
        del msg
        print('Files not found for ',user_name)
        continue
s.quit()
wb.close()






