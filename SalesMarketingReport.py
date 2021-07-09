import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import cx_Oracle
import openpyxl
import os
import sys
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from cryptography.fernet import Fernet
import pandas as pd
import datetime
datevalue=datetime.datetime.today().strftime('%Y%m%d')
timestamp = datetime.datetime.today().strftime('%Y-%m-%d_%H_%M_%S')
print('*****************************************')
print('Script Run DateTime: ',timestamp)
print('Program Execution started....')
# Change current working dir
wrk_dir=r'E:\Control-M\tabprd\FPM\Tableau_Python_Sales_Marketing_Report\\'
os.chdir(wrk_dir)

if not os.path.exists('Archive'):
    os.makedirs('Archive')


# reading excel
wb = openpyxl.load_workbook('Bursting_Details_SalesMarketing.xlsx')
# mail-server details:
sheet = wb["Server&LoginDetails"]
host = sheet['C4'].value
port = sheet['C5'].value
from_id = sheet['C6'].value
user_mailid = sheet['C8'].value
user_ccid = sheet['C9'].value
bcc_id = sheet['C7'].value
fail_mailid = sheet['C10'].value

#database details:
sheet = wb["DatabaseDetails"]
dbuserid = sheet['B1'].value
login_encrypt_password = sheet['B2'].value
dbhost = sheet['B3'].value
dbport = sheet['B4'].value
dbSID = sheet['B5'].value
dbtns=cx_Oracle.makedsn(dbhost,dbport,dbSID)
key=b'icX0uAcHik9UCEgzTY3jP2_KhbEfXGAZucdSX3sbQMQ='
cipher_suite = Fernet(key)
dbpassword = cipher_suite.decrypt(bytes(login_encrypt_password,'utf-8')).decode('utf-8')


def fileconversion(inputfile):
    df = pd.read_csv(inputfile, dtype='unicode')
    df = df.drop([df.shape[0] - 1,df.shape[0] - 2,df.shape[0] - 3], axis='rows')
    df.drop(['Index'],axis=1,inplace=True)
    df = df.rename(columns={'Union Source': 'Source'})
    cols = df.columns.tolist()
    cols.remove('Measure Values')
    df = df.set_index(cols, drop=True).unstack('Measure Names').reset_index()
    df.columns = [''.join(col).strip().replace('Measure Values', '') for col in df.columns.values]
    df['Transactional Amount'] = pd.to_numeric(df['Transactional Amount'].str.replace(",", ""))
    df['Functional Amount'] = pd.to_numeric(df['Functional Amount'].str.replace(",", ""))
    df['Reporting USD Amount'] = pd.to_numeric(df['Reporting USD Amount'].str.replace(",", ""))
    for i in ['Business Unit Id','Account','Department','Product','Business Line Id','Affiliate']:
        df[i] = df[i].apply(lambda x: '{0:0>5}'.format(x) if x is None else x)

    xls_writer = pd.ExcelWriter("output.xlsx", engine='xlsxwriter', datetime_format='dd-mmmm-yyyy')
    cols=['Source','Period Name',	'Ledger',	'Posted Date',	'Segment ID',	'Total Bus Unit ID (TBU)',	'Major Bus Unit ID (MBU)',	'Business Unit Id',	'Business Unit Description',	'Account',	'Account Description',	'Department',	'Product',	'Product Description',	'Business Line Id',	'Affiliate',	'Sub Account',	'Sub Account Description',	'JE Source',	'JE Category',	'JE Number',	'JE Batch Name',	'JE Name',	'JE Description',	'JE Line Number',	'JE Line Description',	'Financial Analyst Name',	'Dept Manager Name',	'Director Name',	'VP Name',	'JE Currency Code',	'AP Vendor Number',	'AP Vendor Name',	'AP Invoice Number',	'AP Invoice Description',	'AP Invoice Line Number',	'AP Invoice Line Descr',	'PO Number',	'PO Requestor Name',	'PO Line Number',	'PO Line Description',	'Po Line Creation Date',	'Po Line Status',	'Acct Category',	'Acct Sub Category',	'Department Name',	'Budget Owner',	'Budget Mgr',	'Business Unit',	'Business Line',	'Mkt - CMO Proirity',	'Transactional Amount',	'Functional Amount',	'Reporting USD Amount',]
    df.to_excel(xls_writer, sheet_name='Sales_Markeing_Report', float_format="%.2f", index=False,columns=cols)
    print('Saving Intermediate file - output.xlsx')
    xls_writer.save()
    print('Opening & Processing Intermediate file - output.xlsx')
    wb = openpyxl.load_workbook("output.xlsx")
    worksheet = wb.active
    worksheet.title = 'Sales_Markeing_Report'

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))


    for col in worksheet.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            cell.font= Font(name='Tahoma',size=8,color='00000000')
            cell.border = thin_border
            # if column in ['AY', 'AZ', 'BA']:
            #     cell.alignment = Alignment(horizontal='left')
            #     cell.value = str(cell.value)

            if cell.coordinate in worksheet.merged_cells:  # not check merge_cells
                continue
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length+2)
        worksheet.column_dimensions[column].width = adjusted_width


    for cell in worksheet["1:1"]:
        cell.font = Font(name='Tahoma',size=8,color='00000000', bold=True)
        cell.fill = PatternFill(start_color="BFD2E2", end_color="FFC7CE", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    c = worksheet['A2']
    worksheet.freeze_panes = c
    print('Saving Final file -',filename)
    wb.save(filename)
    wb.close()
    os.remove("output.xlsx")
    os.remove(inputfile)
def sendErrormail():
    s = smtplib.SMTP(host=host, port=port)
    s.starttls()
    msg = MIMEMultipart()
    msg['From'] = from_id
    msg['To'] = fail_mailid
    msg['Subject'] = 'Action Required: Delivery Failed - Sales and Marketing Report'
    message = 'Error Occurred !'
    msg.attach(MIMEText(message, 'plain'))

    try:
        s.send_message(msg)
        print("Error Occurred Email successfully sent to Internal team")
    except:
        print("Mail Server Error - Error Occurred Email not sent to Internal team")
    del msg
    s.quit()
    sys.exit(0)

def sendFilemail():
    s = smtplib.SMTP(host=host, port=port)
    s.starttls()
    msg = MIMEMultipart()
    msg['From'] = from_id
    msg['To'] = user_mailid
    msg['Cc'] = user_ccid
    msg['Bcc'] = bcc_id
    msg['Subject'] = 'Tableau - Sales and Marketing Report'
    message = '''Hello,
    
Here is the Sales and Marketing Report of Current fiscal year for Current OL version.

For any support issues or questions please contact: bi.help

Thank You,
FPM Team'''
    msg.attach(MIMEText(message, 'plain'))

    mail_attach = MIMEApplication(open(filename, "rb").read())
    mail_attach.add_header('Content-Disposition', 'attachment',
                           filename=filename)
    msg.attach(mail_attach)

    try:
        s.send_message(msg)
        print("Email successfully sent to :",user_mailid)
    except:
        print("Mail Server Error - Email not sent to users")
    del msg
    s.quit()

RunFlag = 'N'
FileFlag='N'
filtercondition = ''
try:
    connection = cx_Oracle.connect(dbuserid,dbpassword,dbtns)
    cursor = connection.cursor()
    querystring = '''select FIS_YEAR from d_calendar where DAY_NAME = 'Friday' and FIS_WK_OF_MONTH = 1 and trunc(NK_CALENDAR_ID) = trunc(sysdate)'''
    cursor.execute(querystring)
    row=cursor.fetchone()
    if row is not None:
        RunFlag = 'Y'
        filtercondition = '?Fiscal%20Year=' + str(row[0])
    cursor.close ()
    connection.close ()

    print('RunFlag - '+ RunFlag)
    FileFlag='N'
except:
    print('Error Occurred while connecting to database')
    sendErrormail()

try:
    if RunFlag =='Y':
        print('File Generation Process started ...')
        # reading Workbook & Dashboard name
        sheet = wb["DashboardNameCoded"]
        workbook = sheet['C2'].value
        listreport = sheet['A2'].value
        # server & log in details:
        sheet = wb["Server&LoginDetails"]
        server = sheet['C1'].value
        login_id = sheet['C2'].value
        login_encrypt_password = sheet['C3'].value
        key = b'icX0uAcHik9UCEgzTY3jP2_KhbEfXGAZucdSX3sbQMQ='
        cipher_suite = Fernet(key)
        login_password = cipher_suite.decrypt(bytes(login_encrypt_password, 'utf-8')).decode('utf-8')
        cmd1 = 'tabcmd login -s ' + server + ' -t SalesandMarketing -u ' + login_id + ' -p ' + login_password + ' --no-certcheck'
        print(os.system(cmd1))

        # file generation Process
        url = '/views/' + workbook + '/' + listreport + '.csv' + filtercondition
        print('URL : ' + url)
        output_csv = 'SalesandMarketingReport.csv'
        cmd2 = r'tabcmd get "' + url + '" -f "' + output_csv + '"'
        print(os.system(cmd2))

        cmd3 = r' tabcmd logout'
        print(os.system(cmd3))

        filename = 'Sales_Markeing_Report_'+datevalue+'.xlsx'
        fileconversion(output_csv)
        FileFlag = 'Y'
        print('File Generated Successfully...')
        if RunFlag == 'Y' and FileFlag == 'Y':
            sendFilemail()
            os.rename(filename, 'Archive\\' + filename[:-5] + '_' + timestamp + '.xlsx')
        else:
            print('Flag are not set properly')
            sendErrormail()
    else:
        print('Today is not 1st Friday of Fiscal Month - No File Generation')
except Exception as e:
    print('Error Occurred in Python Code:',str(e))
    sendErrormail()

