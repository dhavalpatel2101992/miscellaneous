import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from PyPDF2 import PdfFileReader, PdfFileWriter
import time
import openpyxl
import csv
import os
from pyexcel.cookbook import merge_all_to_a_book
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import glob
from openpyxl.styles import PatternFill
from cryptography.fernet import Fernet
import pandas as pd
import datetime
import sys

# Tabcmd location
sys.path.insert(1,r"C:\Program Files\Tableau\Tableau Server\current_prd\extras\Command Line Utility\tabcmd.exe")
print(sys.path)

timestamp = datetime.datetime.today().strftime('%Y-%m-%d_%H:%M:%S')
print('*****************************************')
print('Script Run DateTime: ',timestamp)

def fileconversion(inputfile, user_name, dateparameter):

    merge_all_to_a_book(glob.glob(inputfile), "output.xlsx")

    xl = pd.ExcelFile("output.xlsx")
    df = xl.parse(inputfile)
    df = df.drop([df.shape[0]-1], axis='rows')
    df = df.rename(columns = {'Adj Time Card Date':'Adj Time Card Dt','Month ID':'Month Id','Name':'Full Name','Measure Values':'Hours'})
    df = df.sort_values(['Task Category', 'ITLT L2', 'ITLT L3', 'Full Name'], ascending=[False, True, True, True])
    xls_writer = pd.ExcelWriter("output.xlsx")
    df.to_excel(xls_writer, sheet_name=dateparameter,
                columns=["ITLT L1", "ITLT L2", "ITLT L3","Supervisor Name", "Employee Num", "Full Name","Email Addr","Month Id",
                         "Time Card Dt", "Adj Time Card Dt", "Task Category", "Task Sub Category","Tasks", "Time Card Cd", "Chg Dept",
                         "Home Dept", "Project", "Hours"], index=False)
    xls_writer.save()

    wb = openpyxl.load_workbook("output.xlsx")
    worksheet = wb.active
    worksheet.title = dateparameter

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))


    for col in worksheet.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            cell.font= Font(name='Tahoma',size=8,color='00000000')
            cell.border = thin_border
            if column in ['O','P']:
                cell.alignment = Alignment(horizontal='left')
                cell.value = str(cell.value)

            if cell.coordinate in worksheet.merged_cells:  # not check merge_cells
                continue
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length+2)
        worksheet.column_dimensions[column].width = adjusted_width


    for cell in worksheet["1:1"]:
        cell.font = Font(name='Tahoma',size=8,color='00000000', bold=True)
        cell.fill = PatternFill(start_color="BFD2E2", end_color="FFC7CE", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    c = worksheet['A2']
    worksheet.freeze_panes = c
    wb.save(user_name + ".xlsx")
    wb.close()
    os.remove("output.xlsx")
    os.remove(inputfile)



# Change current working dir
wrk_dir=r'E:\Control-M\tabprd\FPM\Tableau_Python_OneIT_Timetracking\\'
os.chdir(wrk_dir)

# reading excel
wb = openpyxl.load_workbook('Bursting_Details.xlsx')

# reading Workbook & Parameters & Dashboard names
sheet = wb["DashboardNameCoded"]
workbook = sheet['C2'].value
ParameterL2 = sheet['E2'].value
ParameterL3 = sheet['E3'].value
listreport = sheet['G2'].value
l3report = sheet['G3'].value
l3dashboard = sheet['I2'].value
dashboard = []
for i in range(2, sheet.max_row + 1):
    dashboardname = sheet['A' + str(i)].value
    if dashboardname != None:
        dashboard.append(sheet['A' + str(i)].value)
#this will be first page of PDF
dashboard1 = dashboard[0:1]
#these will be last pages of PDF
dashboard2 = dashboard[1:]


# mail-server details:
sheet = wb["Server&LoginDetails"]

host = sheet['C4'].value
port = sheet['C5'].value
from_id = sheet['C6'].value
bcc_id = sheet['C7'].value
subject = sheet['C8'].value



#errornotificationfuction
def sendErrormail(errormsg,user):
    msg = MIMEMultipart()
    msg['From'] = from_id
    msg['To'] = bcc_id
    msg['Subject'] = 'Action Required:'+user+'- Delivery Failed - OneIT TimeTracking Reports'
    message = timestamp+''': Error Occurred while generating OneIT TimeTracking Reports.
    Error Details: '''+str(errormsg)
    msg.attach(MIMEText(message, 'plain'))

    s = smtplib.SMTP(host=host, port=port)
    s.starttls()
    try:
        s.send_message(msg)
        print("Developers have been notified about Failure...")
    except Exception as e:
        print("Mail Server Error - Error Occurred Email not sent to Developers. "+str(e))
    del msg
    s.quit()


# server & log in details:
sheet = wb["Server&LoginDetails"]
server = sheet['C1'].value
login_id = sheet['C2'].value
login_encrypt_password = sheet['C3'].value
key = b'icX0uAcHik9UCEgzTY3jP2_KhbEfXGAZucdSX3sbQMQ='
cipher_suite = Fernet(key)
login_password = cipher_suite.decrypt(bytes(login_encrypt_password, 'utf-8')).decode('utf-8')
cmd1 = 'tabcmd login -s ' + server + ' -t FPM -u ' + login_id + ' -p ' + login_password
print(os.system(cmd1))


# file generation Process
sheet = wb["BurstingList"]
for i in range(2, sheet.max_row + 1):
    user_name = sheet['A' + str(i)].value
    user_namecoded = user_name.replace(",", "").replace(" ", "%20")
    user_namefilecoded = ''.join(filter(str.isalpha, user_name))


    url = '/views/' + workbook + '/' + l3report + '.csv?' + ParameterL2 + '=' + user_namecoded
    output_csv = user_namefilecoded + '_' + l3report + '.csv'
    cmd2 = r'tabcmd get "' + url + '" -f "' + output_csv + '"'
    print(os.system(cmd2))

    try:
        datacheckdf = pd.read_csv(output_csv) #to check data for user exist or not

        url = '/views/' + workbook + '/' + dashboard1[0] + '.pdf?' + ParameterL2 + '=' + user_namecoded
        file_name = user_namefilecoded + '_' + dashboard1[0] + '.pdf'
        cmd2 = r'tabcmd get "' + url + '" -f "' + file_name + '"'
        print(os.system(cmd2))
        time.sleep(10)
        pdf_writer = PdfFileWriter()
        pdf_reader = PdfFileReader(file_name)
        pdf_writer.addPage(pdf_reader.getPage(0))
        os.remove(file_name)

        with open(output_csv) as f:
            reader = csv.reader(f)
            L3data = []
            dateparameter = []
            for row in reader:
                L3data.append(row[1])
                dateparameter.append(row[0])
        dateparameter = dateparameter[1:2]
        finaldateparameter = dateparameter[0]
        finaldateparameter = finaldateparameter.replace('Week Ending - ','Week Ending-')
        L3data = L3data[1:]
        os.remove(output_csv)

        for l3 in L3data:
            l3urlname = l3.replace(' ', '%20')
            l3urlname = l3urlname.replace(',', '')
            url = '/views/' + workbook + '/' + l3dashboard + '.pdf?' + ParameterL2 + '=' + user_namecoded + '&' + ParameterL3 + '=' + l3urlname
            l3filecoded = ''.join(filter(str.isalpha, l3))
            file_name = user_namefilecoded + '_' + l3filecoded + '_' + l3dashboard + '.pdf'
            cmd2 = r'tabcmd get "' + url + '" -f "' + file_name + '"'
            print(os.system(cmd2))
            time.sleep(10)
            pdf_reader = PdfFileReader(file_name)
            pdf_writer.addPage(pdf_reader.getPage(0))
            os.remove(file_name)

        for j in dashboard2:
            url = '/views/' + workbook + '/' + j + '.pdf?' + ParameterL2 + '=' + user_namecoded
            file_name = user_namefilecoded + '_' + j + '.pdf'
            cmd2 = r'tabcmd get "' + url + '" -f "' + file_name + '"'
            print(os.system(cmd2))
            time.sleep(10)
            pdf_reader = PdfFileReader(file_name)
            pdf_writer.addPage(pdf_reader.getPage(0))
            os.remove(file_name)

        output_path = user_namefilecoded + '.pdf'
        with open(output_path, 'wb') as fh:
            pdf_writer.write(fh)

        url = '/views/' + workbook + '/' + listreport + '.csv?' + ParameterL2 + '=' + user_namecoded
        output_csv = user_namefilecoded + '_' + listreport + '.csv'
        cmd2 = r'tabcmd get "' + url + '" -f "' + output_csv + '"'
        print(os.system(cmd2))
        fileconversion(output_csv, user_namefilecoded, finaldateparameter)

    except Exception as e:
        e = "File not generated for "+user_name+". Issue: " + str(e)
        print('Error Details: '+e)
        print('About to send Failure Notification to Developers...')
        sendErrormail(e,user_name)
        os.remove(output_csv)


cmd3 = r' tabcmd logout'
print(os.system(cmd3))
timestamp = datetime.datetime.today().strftime('%Y-%m-%d_%H:%M:%S')
print('Script End DateTime: ',timestamp)
print('*****************************************')
