import cx_Oracle
import pandas as pd
import os
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

datevalue=datetime.datetime.today().strftime('%Y%m%d')
timestamp = datetime.datetime.today().strftime('%Y-%m-%d_%H:%M:%S')
print('*****************************************')
print('Script Run DateTime: ',timestamp)


# Change current working dir
wrk_dir=r'E:\Control-M\tabprd\FPM\Python_Scripts\\'
os.chdir(wrk_dir)

# reading excel
wb = openpyxl.load_workbook('Database_Details.xlsx')
sheet = wb["DatabaseDetails"]
dbuserid = sheet['B1'].value
dbpassword = sheet['B2'].value
dbhost = sheet['B3'].value
dbport = sheet['B4'].value
dbSID = sheet['B5'].value
dbtns=cx_Oracle.makedsn(dbhost,dbport,dbSID)
filepath = sheet['B6'].value
filename = sheet['B7'].value
filename = filepath+'\\'+filename+'_'+datevalue+'.xlsx'
from_id = sheet['B8'].value
fail_mailid = sheet['B9'].value
user_tomailid = sheet['B10'].value
user_ccmailid = sheet['B11'].value
bcc_id=fail_mailid
subject=sheet['B12'].value
mailbody=sheet['B13'].value

s = smtplib.SMTP(host='smtphost.qualcomm.com', port=25)
s.starttls()

def fileformatting(inputfile,sheet):
    wb = openpyxl.load_workbook(inputfile)

    for sh in sheet:
        worksheet = wb[sh]
        thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                cell.font= Font(name='Tahoma',size=8,color='00000000')
                cell.border = thin_border
                # if column in ['J', 'L', 'M']:
                #     cell.alignment = Alignment(horizontal='left')
                #     cell.value = str(cell.value)

                if cell.coordinate in worksheet.merged_cells:  # not check merge_cells
                    continue
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length+2)
            worksheet.column_dimensions[column].width = adjusted_width
        for cell in worksheet["1:1"]:
            cell.font = Font(name='Tahoma',size=8,color='00000000', bold=True)
            cell.fill = PatternFill(start_color="BFD2E2", end_color="FFC7CE", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        c = worksheet['A2']
        worksheet.freeze_panes = c
    wb.save(inputfile)
    wb.close()
def sendErrormail(errormsg):
    msg = MIMEMultipart()
    msg['From'] = from_id
    msg['To'] = fail_mailid
    msg['Subject'] = 'Action Required: Delivery Failed - FCPA New Company Report V3'
    message = timestamp+''': Error Occurred while generating FCPA New Company Report V3 in Python Code.
    Error Details:'''+str(errormsg)
    msg.attach(MIMEText(message, 'plain'))

    try:
        s.send_message(msg)
        print("Developers have been notified about Failure...")
    except:
        print("Mail Server Error - Error Occurred Email not sent to Internal team")
    del msg
def sendnotificationmail():
    msg = MIMEMultipart()
    msg['From'] = from_id
    msg['To'] = user_tomailid
    msg['Cc'] = user_ccmailid
    msg['Bcc'] = bcc_id
    msg['Subject'] = subject
    message = mailbody
    msg.attach(MIMEText(message, 'html'))

    try:
        s.send_message(msg)
        print("Users have been notified Successfully...")
    except Exception as e:
        print("Mail Server Error - Email not sent to users")
        print('Error Details:', e)
        print('About to send Failure Notification to Developers...')
        sendErrormail(e)
    del msg
def senddevelopernotificationmail():
    msg = MIMEMultipart()
    msg['From'] = from_id
    msg['To'] = fail_mailid
    # msg['Cc'] = user_ccmailid
    msg['Bcc'] = bcc_id
    msg['Subject'] = 'FCPA_NewCompanyReport Generated Successfully'
    message =timestamp+ ''': FCPA_NewCompanyReport has been placed successfully at \\canister\cmsdev\SSRS\Deliveries\FCPA'''
    msg.attach(MIMEText(message, 'plain'))

    try:
        s.send_message(msg)
        print("Developers have been notified Successfully...")
    except Exception as e:
        print("Mail Server Error - Email not sent to Developers")
        print('Error Details:', e)
        print('About to send Failure Notification to Developers...')
        sendErrormail(e)
    del msg

try:

    tabs=['New Company Data', 'Additional Data', 'Additional Party Data']

    connection = cx_Oracle.connect(dbuserid,dbpassword,dbtns)
    cursor = connection.cursor()
    querystring = '''
    select COMPANYNAME,LOCATIONCOUNTRY,LOCATIONWEBADDRESS,CREATEDBY,LDW_REFRESH_TIME from LDW_STAGING.S_NVTS_FCPA_NEW_COMPANY_1
    '''
    cursor.execute(querystring)
    df1 = pd.DataFrame(cursor.fetchall())
    querystring = '''
    select COMPANYNAME,LOCATIONCOUNTRY,LOCATIONWEBADDRESS,COMPANYCREATEDBY,COMPANYCREATEDON,CONTRACTNUMBER,ADDITIONALPARTIES,CONTRACTGROUPNAME,
    CONTRACTSTATUSNAME,CONTRACTMANAGER,CONTRACTMANAGERDEPARTMENT,CONTRACTMANAGEREMAIL,CONTRACTTYPEDISPLAY,STRATEGICFUNDINCENTIVE,
    CONTRACTDESCRIPTION,REQUESTOR,REQUESTORLOGIN,REQUESTORDEPARTMENT,REQUESTOREMAIL,REQUESTOR_EMPLOYEE_STATUS,REQUESTOR_SUPERVISOR_NAME,
    REQ_SUPERVISOR_EMAIL_ADDR,CONTRACT_CREATED_BY,CONTRACT_ACTIVATION_DATE,LDW_REFRESH_TIME  from LDW_STAGING.S_NVTS_FCPA_NEW_COMPANY_2
    '''
    cursor.execute(querystring)
    df2 = pd.DataFrame(cursor.fetchall())
    querystring = '''
    select ADDI_PARTY_NAME,ADDI_PARTY_COUNTRY,ADDI_PARTY_LOCATIONWEBADDRESS,ADDI_PARTY_CREATED_BY,ADDI_PARTY_CREATED_ON,CONTRACTNUMBER,
    PRIMARYCOMPANYNAME,CONTRACTADDITIONALPARTIES,CONTRACTGROUPNAME,CONTRACTSTATUSNAME,CONTRACTMANAGER,CONTRACTMANAGERDEPARTMENT,CONTRACTMANAGEREMAIL,
    CONTRACTTYPEDISPLAY,STRATEGICFUNDINCENTIVE,CONTRACTDESCRIPTION,REQUESTOR,REQUESTORLOGIN,REQUESTORDEPARTMENT,REQUESTOREMAIL,REQUESTOR_EMPLOYEE_STATUS,
    REQUESTOR_SUPERVISOR_NAME,REQ_SUPERVISOR_EMAIL_ADDR,CONTRACT_ACTIVATED_BY,CONTRACT_ACTIVATION_DATE,LDW_REFRESH_TIME from LDW_STAGING.S_NVTS_FCPA_NEW_COMPANY_3
    '''
    cursor.execute(querystring)
    df3 = pd.DataFrame(cursor.fetchall())
    #cursor.close () ##commenting line since CLOB error
    #connection.close ()


    xls_writer = pd.ExcelWriter(filename, engine='xlsxwriter',datetime_format='dd-mmmm-yyyy hh:mm:ss',date_format='dd-mmmm-yyyy')
    if (not df1.empty):
        df1 = df1.rename(columns={0: 'Company Name', 1: 'Country', 2: 'WEB Address', 3: 'Company Created By', 4: 'Data Refreshed On'})
        df1 = df1.sort_values(['Company Name', 'Country'], ascending=[True, True])
    else:
        df1=pd.DataFrame(columns=['No Data Available'])
    df1.to_excel(xls_writer, sheet_name=tabs[0], index=False)

    if(not df2.empty):
        df2 = df2.rename(columns={0: 'Company Name',	1: 'Country',	2: 'WEB Address',	3: 'Company Created By',	4: 'Company Created Date',	5: 'Contract Number',	6: 'Contract Additional Parties',	7: 'Contract Group',	8: 'Contract Status',	9: 'Contract Manager',	10: 'Contract Manager Department',	11: 'Contract Manager Email',	12: 'Contract Type',	13: 'Strategic Fund Incentive',	14: 'Contract Description',	15: 'Requestor Name',	16: 'Requestor Login',	17: 'Requestor Department',	18: 'Requestor Email',	19: 'Requestor Employee Status',	20: 'Requestor Supervisor Name',	21: 'Requestor Supervisor Email',	22: 'Contract Activated By',	23: 'Contract Activated Date', 24: 'Data Refreshed On'})
        df2['Company Created Date'] = pd.to_datetime(df2['Company Created Date'], format='%m/%d/%Y')
        df2['Company Created Date'] = df2['Company Created Date'].dt.date
        df2['Contract Activated Date'] = pd.to_datetime(df2['Contract Activated Date'], format='%m/%d/%Y %H:%M:%S')
        df2 = df2.sort_values(['Company Name', 'Country'], ascending=[True, True])
    else:
        df2 = pd.DataFrame(columns=['No Data Available'])
    df2.to_excel(xls_writer, sheet_name=tabs[1], index=False)

    if(not df3.empty):
        df3 = df3.rename(columns={0: 'Additional Party Name',	1: 'Additional Party Country',	2: 'Additional Party WEB Address',	3: 'Additional Party Created By',	4: 'Additional Party Created Date',	5: 'Contract Number',	6: 'Primary Company Name',	7: 'Contract Additional Parties',	8: 'Contract Group',	9: 'Contract Status',	10: 'Contract Manager',	11: 'Contract Manager Department',	12: 'Contract Manager Email',	13: 'Contract Type',	14: 'Strategic Fund Incentive',	15: 'Contract Description',	16: 'Requestor Name',	17: 'Requestor Login',	18: 'Requestor Department',	19: 'Requestor Email',	20: 'Requestor Employee Status',	21: 'Requestor Supervisor Name',	22: 'Requestor Supervisor Email',	23: 'Contract Activated By',	24: 'Contract Activation Date', 25: 'Data Refreshed On'})
        df3['Contract Activation Date'] = pd.to_datetime(df3['Contract Activation Date'], format='%m/%d/%Y %H:%M:%S')
        df3 = df3.sort_values(['Additional Party Name', 'Additional Party Country'], ascending=[True, True])
    else:
        df3 = pd.DataFrame(columns=['No Data Available'])
    df3.to_excel(xls_writer, sheet_name=tabs[2], index=False)
    xls_writer.save()

    fileformatting(filename,tabs)
    print('File generated Successfully...')


    if os.path.isfile(filename):
        print('About to send Report Notification to Developers...')
        senddevelopernotificationmail()
    else:
        print('Error Occurred while generating FCPA New Company Report V3 in Python Code')
        print('Error Details: File Not Found')
        print('About to send Failure Notification to Developers...')
        sendErrormail('File Not Found')

except Exception as e:
    print('Error Occurred while generating FCPA New Company Report V3 in Python Code')
    print('Error Details:',e)
    print('About to send Failure Notification to Developers...')
    sendErrormail(e)


s.quit()
