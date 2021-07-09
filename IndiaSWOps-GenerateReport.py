#!/usr/bin/env python
# coding: utf-8

# In[40]:


import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from PyPDF2 import PdfFileReader, PdfFileWriter
import time
import openpyxl
import csv
import os
from pyexcel.cookbook import merge_all_to_a_book
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import glob
from openpyxl.styles import PatternFill
from cryptography.fernet import Fernet
import pandas as pd
import datetime
import sys
from urllib.parse import quote

# Tabcmd location
sys.path.insert(1,r"C:\Program Files\Tableau\Tableau Server\current_prd\extras\Command Line Utility\tabcmd.exe")

timestamp = datetime.datetime.today().strftime('%Y-%m-%d_%H:%M:%S')
print('*****************************************')
print('Script Run DateTime: ',timestamp)
# Change current working dir
wrk_dir=r'E:\Control-M\tabdev\FPM\Tableau_Python_IndiaSWOps\\'
os.chdir(wrk_dir)
print('Dir: '+wrk_dir)

#errornotificationfuction
def sendErrormail(errormsg,user):
    msg = MIMEMultipart()
    msg['From'] = from_id
    msg['To'] = bcc_id
    msg['Subject'] = 'Action Required:'+user+'- Delivery Failed - India SW Ops Reports'
    message = timestamp+''': Error Occurred while generating India SW Ops Reports.
    Error Details: '''+str(errormsg)
    msg.attach(MIMEText(message, 'plain'))
    s = smtplib.SMTP(host=host, port=port)
    s.starttls()
    try:
        s.send_message(msg)
        print("Developers have been notified about Failure...")
    except Exception as e:
        print("Mail Server Error - Error Occurred Email not sent to Developers. "+str(e))
    del msg
    s.quit()
    
# reading excel
    wb = openpyxl.load_workbook('Bursting_Details_IndiaWSOps.xlsx')

# reading Workbook & Parameters & Dashboard names
sheet = wb["DashboardNameCoded"]
workbook = sheet['C2'].value
parameter = sheet['E2'].value
dashboards = []
for i in range(2, sheet.max_row + 1):
    value = sheet['A' + str(i)].value
    if value != None:
        dashboards.append(value)
        
# mail-server details:
sheet = wb["Server&LoginDetails"]
host = sheet['C4'].value
port = sheet['C5'].value
from_id = sheet['C6'].value
bcc_id = sheet['C7'].value
subject = sheet['C8'].value

# server & log in details:
sheet = wb["Server&LoginDetails"]
server = sheet['C1'].value
login_id = sheet['C2'].value
login_encrypt_password = sheet['C3'].value
key = b'icX0uAcHik9UCEgzTY3jP2_KhbEfXGAZucdSX3sbQMQ='
cipher_suite = Fernet(key)
login_password = cipher_suite.decrypt(bytes(login_encrypt_password, 'utf-8')).decode('utf-8')
cmd1 = 'tabcmd login -s ' + server + ' -t FPM -u ' + login_id + ' -p ' +login_password
if(os.system(cmd1)==0):
    print('Successfully Tableau login : '+server)
else:
    time.sleep(30)
    if(os.system(cmd1)==0):
        print('Successfully login (2nd try): '+server)
    else:
        e='Login process failed'
        print('Error Details: '+e)
        print('About to send Failure Notification to Developers...')
        sendErrormail(e,'All Reports')
        exit(1)
# file generation Process
sheet = wb["BurstingList"]
for i in range(2, sheet.max_row + 1):
    report_view = sheet['A' + str(i)].value
    report_view_coded = quote(report_view, safe='')
    report_view_filename = ''.join(filter(str.isalpha, report_view_coded))
    print('Processing : '+report_view)
    for dashboard in dashboards:
        file_name=report_view_filename+'_'+dashboard + '.png'
        url = '/views/{0}/{1}.png?{2}={3}'.format(workbook,dashboard,parameter,report_view_coded)
        cmd2 = r'tabcmd get "{0}" -f "{1}"'.format(url,file_name)
        print('Command Fired : '+cmd2)
        if(os.system(cmd2)!=0):
            e = "File not generated for " + report_view
            print('Error Details: '+e)
            print('About to send Failure Notification to Developers...')
            sendErrormail(e,report_view)
        time.sleep(10)
        
cmd3 = r' tabcmd logout'
if(os.system(cmd3)==0):
    print('Successfully Tableau logout : '+server)
timestamp = datetime.datetime.today().strftime('%Y-%m-%d_%H:%M:%S')
print('Script End DateTime: ',timestamp)
print('*****************************************')

