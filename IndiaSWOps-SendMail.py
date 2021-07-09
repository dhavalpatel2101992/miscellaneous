#!/usr/bin/env python
# coding: utf-8

# In[16]:


import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from PyPDF2 import PdfFileReader, PdfFileWriter
import time
import openpyxl
import csv
import os
from pyexcel.cookbook import merge_all_to_a_book
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import glob
from openpyxl.styles import PatternFill
from cryptography.fernet import Fernet
import pandas as pd
import datetime
import sys
import cv2
from urllib.parse import quote

timestamp = datetime.datetime.today().strftime('%Y-%m-%d_%H:%M:%S')
print('*****************************************')
print('Script Run DateTime: ',timestamp)

# Change current working dir
wrk_dir=r'E:\Control-M\tabprd\FPM\Tableau_Python_IndiaSWOps\\'
os.chdir(wrk_dir)
print('Dir: '+wrk_dir)

# Image processing
def remove_whitespace(file_name):
    image=cv2.imread(file_name)
    gray=cv2.cvtColor(image,cv2.COLOR_BGR2GRAY)
    _,binary=cv2.threshold(gray,0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)
    mask=cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (151,151))
    morphed=cv2.morphologyEx(binary, cv2.MORPH_CLOSE, mask)
    cnts,_= cv2.findContours(morphed.copy(),cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts= sorted(cnts, key=cv2.contourArea, reverse= True)
    if len(cnts)!=0:
        c = max(cnts, key = cv2.contourArea)
        x,y,w,h = cv2.boundingRect(c)
        try:
            image = image[0:y + h + 10, 0:x + w]
        except:
            image = image[0:y + h, 0:x + w]
    cv2.imwrite(file_name,image)
    print("{0} Processed Successfully ...".format(file_name))
    return y+h,x+w

#errornotificationfuction
def sendErrormail(errormsg,user):
    msg = MIMEMultipart()
    msg['From'] = from_id
    msg['To'] = bcc_id
    msg['Subject'] = 'Action Required:'+user+'- Delivery Failed - India SW Ops Reports'
    message = timestamp+''': Error Occurred while generating India SW Ops Reports.
    Error Details: '''+str(errormsg)
    msg.attach(MIMEText(message, 'plain'))
    s = smtplib.SMTP(host=host, port=port)
    s.starttls()
    try:
        s.send_message(msg)
        print("Developers have been notified about Failure...")
    except Exception as e:
        print("Mail Server Error - Error Occurred Email not sent to Developers. "+str(e))
    del msg
    s.quit()

# reading excel
wb = openpyxl.load_workbook('Bursting_Details.xlsx')

# reading Dashboard names
sheet = wb["DashboardNameCoded"]
workbook = sheet['C2'].value
parameter = sheet['E2'].value
dashboards = []
for i in range(2, sheet.max_row + 1):
    value = sheet['A' + str(i)].value
    if value != None:
        dashboards.append(value)

# mail-server details
sheet = wb["Server&LoginDetails"]
server = sheet['C1'].value
host = sheet['C4'].value
port = sheet['C5'].value
from_id = sheet['C6'].value
bcc_id = sheet['C7'].value
subject = sheet['C8'].value

# Bursting Process
sheet = wb["BurstingList"]
for i in range(2, sheet.max_row + 1):
    report_view = sheet['A' + str(i)].value
    print('Processing : '+report_view)
    msg = MIMEMultipart()
    msg['From'] = from_id
    msg['Subject'] = subject + report_view
    msg['Bcc'] = bcc_id
    if datetime.datetime.today().weekday() == 2:
        user_mailid = sheet['C' + str(i)].value
        msg['To'] = user_mailid
        print('Today is Wednesday Hence column C in use in TO')
    else:
        user_mailid = sheet['B' + str(i)].value
        msg['To'] = user_mailid
        print('Today is not Wednesday Hence column B in use in TO')
    print('To ID:{0}'.format(msg['To']))
    print('Bcc ID:{0}'.format(msg['Bcc']))
    message = '<html> <body>'
    workbookurl = False
    report_view_coded = quote(report_view, safe='')
    report_view_filename = ''.join(filter(str.isalpha, report_view_coded))
    for dashboard in dashboards:
        try:
            file_name=report_view_filename+'_'+dashboard + '.png'
            h, w = remove_whitespace(file_name)
            viewurl = '{0}/#/site/FPM/views/{1}/{2}?{3}={4}'.format(server, workbook, dashboard,parameter,report_view_coded)
            if not workbookurl:
                workbookurl = viewurl
            tag = '<a href="' + viewurl + '"> <img src="cid:' + file_name + '" width={0} height={1}> </a>'.format(w, h)
            message = message + tag
            fp = open(file_name, 'rb')
            msgImage = MIMEImage(fp.read())
            fp.close()
            imtg = '<' + file_name + '>'
            msgImage.add_header('Content-ID', imtg)
            msg.attach(msgImage)
            timestamp = datetime.datetime.today().strftime('%Y-%m-%d_%H-%M-%S')
            os.rename(file_name, 'Archive\\' + file_name.replace('.png','') + '_' + timestamp + '.png')
        except Exception as e:
            e = "File not found -  " + file_name + ": Issue: " + str(e)
            print('Error Details: ' + e)
            print('About to send Failure Notification to Developers...')
            sendErrormail(e,report_view)
            workbookurl=False
            break
    if workbookurl:
        footer = '''
        <br><br>
        ***This is an automated email message. For any query or require further information please contact indiasw.hc@qti.qualcomm.com
        <br><br>
        Here's your subscription to the workbook <a href="{0}">India SW Ops - {1}</a>
        </body></html>
        '''.format(workbookurl, report_view)
        message = message + footer
        msg.attach(MIMEText(message, 'html'))
        s = smtplib.SMTP(host=host, port=port)
        s.starttls()
        try:
            s.send_message(msg)
            print("Email successfully sent to",user_mailid)
        except:
            print("Error - Email not sent to",user_mailid)
        del msg
        s.quit()
timestamp = datetime.datetime.today().strftime('%Y-%m-%d_%H:%M:%S')
print('Script End DateTime: ',timestamp)
print('*****************************************')

